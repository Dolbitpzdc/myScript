import openpyxl


f2 = open('product.txt', 'w')
f2.write('')
f2.close()


wb = openpyxl.reader.excel.load_workbook(filename= "price.xlsx", data_only=True)

print(wb.sheetnames)

wb.active = 0

sheet = wb.active

f = open('product.txt', 'r+')

for i in range(485, 499):


    filename = str(sheet['C' + str(i)].value)
    product = str(sheet['E' + str(i)].value)
    price = float(sheet['F' + str(i)].value)


    f.write('''<div class="col-md-3 col-6">
      <div class="product-wrap border">
        <div class="product-item ">
          <img src="{% static 'img/disk.jpg' %}">
          <div class="product-buttons">
            <a href="" class="button">Перейти</a>
          </div>
        </div>
        <div class="product-title">
          <a href="">''' +filename+ '''</a>
        </div>
        <div class="product-price1">Миним заказ, шт: '''+product+'''</div>
        <span class="product-price">'''+str(int(price))+''' руб/шт</span>
      </div>
    </div>''' + "\n")



f.close()




# print(sheet['C' + str(i)].value, sheet['E' + str(i)].value, sheet['F' + str(i)].value)
# sheet = book.active
#
# # cells = sheet['C665:F696']
#
# # for row in sheet.iter_rows(min_row=665, max_row=696, min_col=3, max_col=5):
# #     for cell in row:
# #         print(cell.value, end='  ')
# #     print()
#
# # for name, col, price in cells:
# #
# #     print(name.value, col.value, price.value)
#
# for row in range(5,sheet.row[665], sheet.max_row+1):
#     name = sheet[row][2].value
#     col = sheet[row][6].value
#     price = sheet[row][7].value
#
#     print(row, name, col, price)
